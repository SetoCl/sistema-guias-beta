import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, "database.db")
ARCHIVO = os.path.join(BASE_DIR, "datos.xlsx")


def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def formatear_fecha(fecha_excel):
    if not fecha_excel:
        return None

    if isinstance(fecha_excel, datetime):
        return fecha_excel.strftime("%Y-%m-%d")

    try:
        return str(fecha_excel)
    except:
        return None


def importar():

    if not os.path.exists(ARCHIVO):
        print("❌ No se encontró datos.xlsx")
        return

    wb = load_workbook(ARCHIVO, read_only=True)
    ws = wb["guias"]

    conn = get_db_connection()

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"Filas detectadas: {len(filas)}")

    contador = 0

    for fila in filas:

        if not any(fila):
            continue

        numero_guia = fila[0]
        fecha = formatear_fecha(fila[1])
        edificio = fila[2]
        direccion = fila[3]
        property_val = fila[4]
        descripcion = fila[5]

        tecnico1 = fila[6]
        tecnico2 = fila[7]
        tecnico3 = fila[8]
        tecnico4 = fila[9]
        tecnico5 = fila[10]
        tecnico6 = fila[11]

        observaciones = fila[12]

        if not numero_guia:
            continue

        existe = conn.execute(
            "SELECT id FROM guias WHERE numero_guia = ?",
            (numero_guia,)
        ).fetchone()

        if existe:
            continue

        lista_tecnicos = [
            t for t in [
                tecnico1, tecnico2, tecnico3,
                tecnico4, tecnico5, tecnico6
            ] if t
        ]

        tecnicos = ", ".join(lista_tecnicos)
        tecnico_firma = tecnico1 if tecnico1 else ""

        # 🔵 Cliente completo
        cliente = conn.execute(
            "SELECT id FROM clientes WHERE nombre = ?",
            (edificio,)
        ).fetchone()

        if not cliente:
            conn.execute(
                "INSERT INTO clientes (nombre, direccion, property) VALUES (?, ?, ?)",
                (edificio, direccion or "", property_val or "")
            )
            cliente = conn.execute(
                "SELECT id FROM clientes WHERE nombre = ?",
                (edificio,)
            ).fetchone()

        cliente_id = cliente["id"]

        conn.execute("""
            INSERT INTO guias
            (numero_guia, fecha, cliente_id, tecnicos,
             tecnico_firma, descripcion, estado, observaciones)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            numero_guia,
            fecha,
            cliente_id,
            tecnicos,
            tecnico_firma,
            descripcion,
            "Cerrada",
            observaciones
        ))

        contador += 1

        if contador % 50 == 0:
            print(f"Procesadas: {contador}")

    conn.commit()
    conn.close()

    print(f"✅ Importación completada. Guías cargadas: {contador}")


if __name__ == "__main__":
    importar()