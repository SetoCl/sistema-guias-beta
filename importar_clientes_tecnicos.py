import sqlite3
import os
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE = os.path.join(BASE_DIR, "database.db")
ARCHIVO = os.path.join(BASE_DIR, "datos.xlsx")


def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def importar():

    if not os.path.exists(ARCHIVO):
        print("❌ No se encontró datos.xlsx")
        return

    print("Abriendo Excel...")
    wb = load_workbook(ARCHIVO, read_only=True)

    conn = get_db_connection()

    clientes_creados = 0
    tecnicos_creados = 0

    # ============================
    # CLIENTES
    # ============================
    if "clientes" in wb.sheetnames:
        ws_clientes = wb["clientes"]

        print("Importando clientes...")

        for i, fila in enumerate(ws_clientes.iter_rows(min_row=2, values_only=True)):

            if not any(fila):
                continue

            nombre = fila[0]
            property = fila[1]
            direccion = fila[2]

            if not nombre:
                continue

            existe = conn.execute(
                "SELECT id FROM clientes WHERE nombre = ?",
                (nombre,)
            ).fetchone()

            if not existe:
                conn.execute(
                    "INSERT INTO clientes (nombre, direccion, property) VALUES (?, ?, ?)",
                    (nombre, direccion or "", property or "")
                )
                clientes_creados += 1

            if i % 100 == 0:
                print(f"Procesados clientes: {i}")

    else:
        print("⚠ No existe hoja 'clientes'")

    # ============================
    # TECNICOS
    # ============================
    if "tecnicos" in wb.sheetnames:
        ws_tecnicos = wb["tecnicos"]

        print("Importando técnicos...")

        for i, fila in enumerate(ws_tecnicos.iter_rows(min_row=2, values_only=True)):

            if not any(fila):
                continue

            nombre = fila[0]

            if not nombre:
                continue

            existe = conn.execute(
                "SELECT id FROM tecnicos WHERE nombre = ?",
                (nombre,)
            ).fetchone()

            if not existe:
                conn.execute(
                    "INSERT INTO tecnicos (nombre) VALUES (?)",
                    (nombre,)
                )
                tecnicos_creados += 1

            if i % 100 == 0:
                print(f"Procesados técnicos: {i}")

    else:
        print("⚠ No existe hoja 'tecnicos'")

    conn.commit()
    conn.close()

    print("✅ Importación completada")
    print("Clientes creados:", clientes_creados)
    print("Técnicos creados:", tecnicos_creados)


if __name__ == "__main__":
    importar()